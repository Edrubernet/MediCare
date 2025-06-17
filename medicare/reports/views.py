import json
import io
from datetime import datetime, timedelta
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Avg, Q, Sum
from django.http import JsonResponse, HttpResponse, Http404
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import TemplateView, ListView, DetailView, CreateView
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment
import csv

from .models import Report, ReportExport
from .serializers import ReportSerializer
from patients.models import PatientProfile
from programs.models import RehabilitationProgram
from progress.models import ExerciseProgress
from staff.models import User
from consultation.models import Consultation
from survey.models import SurveyResponse


class ReportViewSet(viewsets.ModelViewSet):
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Report.objects.filter(created_by=self.request.user)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class ReportDashboardView(LoginRequiredMixin, TemplateView):
    """Головна панель звітів"""
    template_name = 'reports/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика системи
        context['total_patients'] = PatientProfile.objects.count()
        context['total_programs'] = RehabilitationProgram.objects.count()
        context['total_consultations'] = Consultation.objects.count()
        context['active_programs'] = RehabilitationProgram.objects.filter(is_active=True).count()
        
        # Останні звіти
        context['recent_reports'] = Report.objects.filter(
            created_by=self.request.user
        ).order_by('-created_at')[:5]
        
        # Швидкі звіти
        context['quick_stats'] = self.get_quick_stats()
        
        return context
    
    def get_quick_stats(self):
        """Швидка статистика для дашборда"""
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        return {
            'new_patients_week': PatientProfile.objects.filter(created_at__gte=week_ago).count(),
            'consultations_week': Consultation.objects.filter(created_at__gte=week_ago).count(),
            'completed_exercises_week': ExerciseProgress.objects.filter(
                completed_date__gte=week_ago,
                is_completed=True
            ).count(),
            'active_programs_month': RehabilitationProgram.objects.filter(
                created_at__gte=month_ago,
                is_active=True
            ).count(),
        }


class CreateReportView(LoginRequiredMixin, CreateView):
    """Створення нового звіту"""
    model = Report
    template_name = 'reports/create.html'
    fields = ['title', 'report_type', 'description', 'parameters']
    success_url = reverse_lazy('reports:dashboard')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        
        # Обробка JSON параметрів
        parameters = self.request.POST.get('parameters_json', '{}')
        try:
            form.instance.parameters = json.loads(parameters)
        except json.JSONDecodeError:
            form.instance.parameters = {}
        
        messages.success(self.request, _('Report created successfully.'))
        return super().form_valid(form)


class ReportDetailView(LoginRequiredMixin, DetailView):
    """Деталі звіту"""
    model = Report
    template_name = 'reports/detail.html'
    context_object_name = 'report'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Генеруємо дані звіту
        report_data = self.generate_report_data(self.object)
        context['report_data'] = report_data
        
        # Експорти звіту
        context['exports'] = self.object.exports.all().order_by('-created_at')
        
        return context
    
    def generate_report_data(self, report):
        """Генерує дані для звіту"""
        if report.report_type == 'patient_progress':
            return self.generate_patient_progress_data(report)
        elif report.report_type == 'program_completion':
            return self.generate_program_completion_data(report)
        elif report.report_type == 'doctor_activity':
            return self.generate_doctor_activity_data(report)
        elif report.report_type == 'survey_results':
            return self.generate_survey_results_data(report)
        elif report.report_type == 'system_usage':
            return self.generate_system_usage_data(report)
        else:
            return {}
    
    def generate_patient_progress_data(self, report):
        """Генерує дані прогресу пацієнтів"""
        patients = PatientProfile.objects.all()
        
        data = []
        for patient in patients:
            progress_stats = ExerciseProgress.objects.filter(
                assignment__patient=patient
            ).aggregate(
                total_exercises=Count('id'),
                completed_exercises=Count('id', filter=Q(is_completed=True)),
                avg_score=Avg('score')
            )
            
            completion_rate = 0
            if progress_stats['total_exercises']:
                completion_rate = (progress_stats['completed_exercises'] / progress_stats['total_exercises']) * 100
            
            data.append({
                'patient_name': patient.user.get_full_name(),
                'total_exercises': progress_stats['total_exercises'] or 0,
                'completed_exercises': progress_stats['completed_exercises'] or 0,
                'completion_rate': round(completion_rate, 1),
                'avg_score': round(progress_stats['avg_score'] or 0, 1),
                'active_programs': patient.assigned_programs.filter(is_active=True).count()
            })
        
        return {
            'patients': data,
            'summary': {
                'total_patients': len(data),
                'avg_completion_rate': round(sum(p['completion_rate'] for p in data) / len(data) if data else 0, 1)
            }
        }
    
    def generate_program_completion_data(self, report):
        """Генерує дані завершення програм"""
        programs = RehabilitationProgram.objects.all()
        
        data = []
        for program in programs:
            assignments = program.assignments.all()
            completed_assignments = assignments.filter(status='completed')
            
            completion_rate = 0
            if assignments.exists():
                completion_rate = (completed_assignments.count() / assignments.count()) * 100
            
            data.append({
                'program_title': program.title,
                'total_assignments': assignments.count(),
                'completed_assignments': completed_assignments.count(),
                'completion_rate': round(completion_rate, 1),
                'avg_duration': program.duration_weeks,
                'created_by': program.created_by.get_full_name()
            })
        
        return {
            'programs': data,
            'summary': {
                'total_programs': len(data),
                'avg_completion_rate': round(sum(p['completion_rate'] for p in data) / len(data) if data else 0, 1)
            }
        }
    
    def generate_doctor_activity_data(self, report):
        """Генерує дані активності лікарів"""
        doctors = User.objects.filter(user_type='staff')
        
        data = []
        for doctor in doctors:
            activity_stats = {
                'consultations': Consultation.objects.filter(doctor=doctor).count(),
                'programs_created': RehabilitationProgram.objects.filter(created_by=doctor).count(),
                'patients_assigned': PatientProfile.objects.filter(assigned_doctor=doctor).count(),
            }
            
            data.append({
                'doctor_name': doctor.get_full_name(),
                'consultations': activity_stats['consultations'],
                'programs_created': activity_stats['programs_created'],
                'patients_assigned': activity_stats['patients_assigned'],
                'specialization': getattr(doctor.staff_profile, 'specialization', '') if hasattr(doctor, 'staff_profile') else ''
            })
        
        return {
            'doctors': data,
            'summary': {
                'total_doctors': len(data),
                'total_consultations': sum(d['consultations'] for d in data),
                'total_programs': sum(d['programs_created'] for d in data)
            }
        }
    
    def generate_survey_results_data(self, report):
        """Генерує дані результатів опитувань"""
        responses = SurveyResponse.objects.filter(completed_at__isnull=False)
        
        data = []
        for response in responses:
            data.append({
                'survey_title': response.assignment.survey.title,
                'patient_name': response.assignment.patient.user.get_full_name(),
                'completed_date': response.completed_at.strftime('%Y-%m-%d'),
                'questions_answered': response.question_responses.count(),
                'assigned_by': response.assignment.assigned_by.get_full_name()
            })
        
        return {
            'responses': data,
            'summary': {
                'total_responses': len(data),
                'completion_rate': round((len(data) / max(responses.count(), 1)) * 100, 1)
            }
        }
    
    def generate_system_usage_data(self, report):
        """Генерує дані використання системи"""
        today = datetime.now().date()
        days_30 = today - timedelta(days=30)
        
        usage_data = {
            'user_registrations': {
                'last_30_days': User.objects.filter(date_joined__gte=days_30).count(),
                'total': User.objects.count()
            },
            'consultations': {
                'last_30_days': Consultation.objects.filter(created_at__gte=days_30).count(),
                'total': Consultation.objects.count()
            },
            'exercises_completed': {
                'last_30_days': ExerciseProgress.objects.filter(
                    completed_date__gte=days_30,
                    is_completed=True
                ).count(),
                'total': ExerciseProgress.objects.filter(is_completed=True).count()
            },
            'programs_active': RehabilitationProgram.objects.filter(is_active=True).count()
        }
        
        return usage_data


class PatientProgressReportView(LoginRequiredMixin, TemplateView):
    """Звіт прогресу пацієнтів"""
    template_name = 'reports/patient_progress.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Фільтри
        patient_id = self.request.GET.get('patient_id')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        # Базовий queryset
        progress_queryset = ExerciseProgress.objects.all()
        
        if patient_id:
            progress_queryset = progress_queryset.filter(assignment__patient_id=patient_id)
        
        if date_from:
            progress_queryset = progress_queryset.filter(assignment__assigned_date__gte=date_from)
        
        if date_to:
            progress_queryset = progress_queryset.filter(assignment__assigned_date__lte=date_to)
        
        # Статистика
        context['progress_data'] = progress_queryset.select_related(
            'assignment__patient__user',
            'assignment__program',
            'exercise'
        ).order_by('-completed_date')
        
        context['patients'] = PatientProfile.objects.all()
        context['filters'] = {
            'patient_id': patient_id,
            'date_from': date_from,
            'date_to': date_to
        }
        
        return context


class ProgramCompletionReportView(LoginRequiredMixin, TemplateView):
    """Звіт завершення програм"""
    template_name = 'reports/program_completion.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        programs = RehabilitationProgram.objects.all()
        program_stats = []
        
        for program in programs:
            assignments = program.assignments.all()
            completed = assignments.filter(status='completed').count()
            in_progress = assignments.filter(status='in_progress').count()
            
            program_stats.append({
                'program': program,
                'total_assignments': assignments.count(),
                'completed': completed,
                'in_progress': in_progress,
                'completion_rate': round((completed / assignments.count() * 100) if assignments.count() > 0 else 0, 1)
            })
        
        context['program_stats'] = program_stats
        return context


class DoctorActivityReportView(LoginRequiredMixin, TemplateView):
    """Звіт активності лікарів"""
    template_name = 'reports/doctor_activity.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        doctors = User.objects.filter(user_type='staff')
        doctor_stats = []
        
        for doctor in doctors:
            stats = {
                'doctor': doctor,
                'consultations_count': Consultation.objects.filter(doctor=doctor).count(),
                'programs_created': RehabilitationProgram.objects.filter(created_by=doctor).count(),
                'patients_count': PatientProfile.objects.filter(assigned_doctor=doctor).count(),
                'recent_consultations': Consultation.objects.filter(doctor=doctor).order_by('-created_at')[:5]
            }
            doctor_stats.append(stats)
        
        context['doctor_stats'] = doctor_stats
        return context


class SurveyResultsReportView(LoginRequiredMixin, TemplateView):
    """Звіт результатів опитувань"""
    template_name = 'reports/survey_results.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        responses = SurveyResponse.objects.filter(
            completed_at__isnull=False
        ).select_related(
            'assignment__survey',
            'assignment__patient__user'
        ).order_by('-completed_at')
        
        context['survey_responses'] = responses
        return context


class SystemUsageReportView(LoginRequiredMixin, TemplateView):
    """Звіт використання системи"""
    template_name = 'reports/system_usage.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика за різні періоди
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        usage_stats = {
            'users': {
                'total': User.objects.count(),
                'patients': User.objects.filter(user_type='patient').count(),
                'staff': User.objects.filter(user_type='staff').count(),
                'new_this_week': User.objects.filter(date_joined__gte=week_ago).count(),
                'new_this_month': User.objects.filter(date_joined__gte=month_ago).count(),
            },
            'activity': {
                'consultations_total': Consultation.objects.count(),
                'consultations_week': Consultation.objects.filter(created_at__gte=week_ago).count(),
                'programs_active': RehabilitationProgram.objects.filter(is_active=True).count(),
                'exercises_completed_month': ExerciseProgress.objects.filter(
                    completed_date__gte=month_ago,
                    is_completed=True
                ).count(),
            }
        }
        
        context['usage_stats'] = usage_stats
        return context


class ExportReportView(LoginRequiredMixin, View):
    """Експорт звіту"""
    
    def get(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        export_format = request.GET.get('format', 'pdf')
        
        if export_format == 'pdf':
            return self.export_pdf(report)
        elif export_format == 'csv':
            return self.export_csv(report)
        elif export_format == 'excel':
            return self.export_excel(report)
        else:
            messages.error(request, _('Invalid export format.'))
            return redirect('reports:report_detail', pk=pk)
    
    def export_pdf(self, report):
        """Експорт в PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Заголовок
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Центрування
        )
        story.append(Paragraph(report.title, title_style))
        story.append(Spacer(1, 12))
        
        # Опис
        if report.description:
            story.append(Paragraph(f"<b>Опис:</b> {report.description}", styles['Normal']))
            story.append(Spacer(1, 12))
        
        # Дата створення
        story.append(Paragraph(f"<b>Дата створення:</b> {report.created_at.strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        story.append(Paragraph(f"<b>Створено:</b> {report.created_by.get_full_name()}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Дані звіту
        report_data = self.generate_report_data(report)
        if report_data:
            story.append(Paragraph("<b>Дані звіту:</b>", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Конвертуємо дані в таблицю
            if report.report_type == 'patient_progress' and 'patients' in report_data:
                data = [['Пацієнт', 'Всього вправ', 'Виконано', 'Відсоток', 'Сер. оцінка']]
                for patient in report_data['patients']:
                    data.append([
                        patient['patient_name'],
                        str(patient['total_exercises']),
                        str(patient['completed_exercises']),
                        f"{patient['completion_rate']}%",
                        str(patient['avg_score'])
                    ])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report.title}.pdf"'
        
        # Зберігаємо експорт
        ReportExport.objects.create(
            report=report,
            format='pdf',
            file=f'report_exports/{report.title}.pdf'
        )
        
        return response
    
    def export_csv(self, report):
        """Експорт в CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{report.title}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Звіт:', report.title])
        writer.writerow(['Дата:', report.created_at.strftime('%d.%m.%Y %H:%M')])
        writer.writerow(['Створено:', report.created_by.get_full_name()])
        writer.writerow([])  # Порожній рядок
        
        # Дані звіту
        report_data = self.generate_report_data(report)
        if report.report_type == 'patient_progress' and 'patients' in report_data:
            writer.writerow(['Пацієнт', 'Всього вправ', 'Виконано', 'Відсоток виконання', 'Середня оцінка'])
            for patient in report_data['patients']:
                writer.writerow([
                    patient['patient_name'],
                    patient['total_exercises'],
                    patient['completed_exercises'],
                    f"{patient['completion_rate']}%",
                    patient['avg_score']
                ])
        
        return response
    
    def export_excel(self, report):
        """Експорт в Excel"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Звіт"
        
        # Заголовки
        worksheet['A1'] = report.title
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A3'] = f"Дата створення: {report.created_at.strftime('%d.%m.%Y %H:%M')}"
        worksheet['A4'] = f"Створено: {report.created_by.get_full_name()}"
        
        # Дані звіту
        report_data = self.generate_report_data(report)
        if report.report_type == 'patient_progress' and 'patients' in report_data:
            # Заголовки стовпців
            headers = ['Пацієнт', 'Всього вправ', 'Виконано', 'Відсоток виконання', 'Середня оцінка']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=6, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Дані
            for row, patient in enumerate(report_data['patients'], 7):
                worksheet.cell(row=row, column=1, value=patient['patient_name'])
                worksheet.cell(row=row, column=2, value=patient['total_exercises'])
                worksheet.cell(row=row, column=3, value=patient['completed_exercises'])
                worksheet.cell(row=row, column=4, value=f"{patient['completion_rate']}%")
                worksheet.cell(row=row, column=5, value=patient['avg_score'])
        
        # Зберігаємо в пам'ять
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report.title}.xlsx"'
        
        return response
    
    def generate_report_data(self, report):
        """Генерує дані звіту (дублює логіку з ReportDetailView)"""
        # Тут повторюємо логіку генерації даних
        if report.report_type == 'patient_progress':
            patients = PatientProfile.objects.all()
            data = []
            for patient in patients:
                progress_stats = ExerciseProgress.objects.filter(
                    assignment__patient=patient
                ).aggregate(
                    total_exercises=Count('id'),
                    completed_exercises=Count('id', filter=Q(is_completed=True)),
                    avg_score=Avg('score')
                )
                
                completion_rate = 0
                if progress_stats['total_exercises']:
                    completion_rate = (progress_stats['completed_exercises'] / progress_stats['total_exercises']) * 100
                
                data.append({
                    'patient_name': patient.user.get_full_name(),
                    'total_exercises': progress_stats['total_exercises'] or 0,
                    'completed_exercises': progress_stats['completed_exercises'] or 0,
                    'completion_rate': round(completion_rate, 1),
                    'avg_score': round(progress_stats['avg_score'] or 0, 1)
                })
            
            return {'patients': data}
        
        return {}


# Інші view-класи для аналітики
class PatientAnalyticsView(LoginRequiredMixin, View):
    """API для аналітики конкретного пацієнта"""
    
    def get(self, request, patient_id):
        patient = get_object_or_404(PatientProfile, id=patient_id)
        
        analytics_data = {
            'patient_info': {
                'name': patient.user.get_full_name(),
                'email': patient.user.email,
                'phone': patient.phone_number,
                'registered_date': patient.created_at.strftime('%Y-%m-%d')
            },
            'progress_summary': self.get_patient_progress_summary(patient),
            'programs_history': self.get_patient_programs(patient),
            'consultations_history': self.get_patient_consultations(patient)
        }
        
        return JsonResponse(analytics_data)
    
    def get_patient_progress_summary(self, patient):
        """Зводка прогресу пацієнта"""
        progress = ExerciseProgress.objects.filter(assignment__patient=patient)
        
        return {
            'total_exercises': progress.count(),
            'completed_exercises': progress.filter(is_completed=True).count(),
            'avg_score': round(progress.aggregate(avg_score=Avg('score'))['avg_score'] or 0, 1),
            'completion_rate': round((progress.filter(is_completed=True).count() / progress.count() * 100) if progress.count() > 0 else 0, 1)
        }
    
    def get_patient_programs(self, patient):
        """Історія програм пацієнта"""
        programs = patient.assigned_programs.all()
        
        return [{
            'title': program.title,
            'status': 'active' if program.is_active else 'completed',
            'start_date': program.created_at.strftime('%Y-%m-%d'),
            'duration_weeks': program.duration_weeks
        } for program in programs]
    
    def get_patient_consultations(self, patient):
        """Історія консультацій пацієнта"""
        consultations = Consultation.objects.filter(patient=patient).order_by('-scheduled_date')
        
        return [{
            'date': consultation.scheduled_date.strftime('%Y-%m-%d %H:%M'),
            'doctor': consultation.doctor.get_full_name(),
            'status': consultation.status,
            'notes': consultation.notes or ''
        } for consultation in consultations[:10]]  # Останні 10


class ProgramAnalyticsView(LoginRequiredMixin, View):
    """API для аналітики програми"""
    
    def get(self, request, program_id):
        program = get_object_or_404(RehabilitationProgram, id=program_id)
        
        analytics_data = {
            'program_info': {
                'title': program.title,
                'description': program.description,
                'created_by': program.created_by.get_full_name(),
                'created_date': program.created_at.strftime('%Y-%m-%d'),
                'duration_weeks': program.duration_weeks,
                'is_active': program.is_active
            },
            'assignments_summary': self.get_program_assignments_summary(program),
            'exercises_stats': self.get_program_exercises_stats(program)
        }
        
        return JsonResponse(analytics_data)
    
    def get_program_assignments_summary(self, program):
        """Зводка призначень програми"""
        assignments = program.assignments.all()
        
        return {
            'total_assignments': assignments.count(),
            'completed_assignments': assignments.filter(status='completed').count(),
            'in_progress_assignments': assignments.filter(status='in_progress').count(),
            'completion_rate': round((assignments.filter(status='completed').count() / assignments.count() * 100) if assignments.count() > 0 else 0, 1)
        }
    
    def get_program_exercises_stats(self, program):
        """Статистика вправ програми"""
        exercises = program.exercises.all()
        
        return [{
            'exercise_title': exercise.title,
            'total_attempts': ExerciseProgress.objects.filter(exercise=exercise).count(),
            'completed_attempts': ExerciseProgress.objects.filter(exercise=exercise, is_completed=True).count(),
            'avg_score': round(ExerciseProgress.objects.filter(exercise=exercise).aggregate(avg_score=Avg('score'))['avg_score'] or 0, 1)
        } for exercise in exercises]


class SystemOverviewView(LoginRequiredMixin, View):
    """API для загального огляду системи"""
    
    def get(self, request):
        overview_data = {
            'users_stats': self.get_users_stats(),
            'activity_stats': self.get_activity_stats(),
            'growth_stats': self.get_growth_stats(),
            'performance_stats': self.get_performance_stats()
        }
        
        return JsonResponse(overview_data)
    
    def get_users_stats(self):
        """Статистика користувачів"""
        return {
            'total_users': User.objects.count(),
            'patients': User.objects.filter(user_type='patient').count(),
            'staff': User.objects.filter(user_type='staff').count(),
            'active_users': User.objects.filter(is_active=True).count()
        }
    
    def get_activity_stats(self):
        """Статистика активності"""
        return {
            'total_consultations': Consultation.objects.count(),
            'active_programs': RehabilitationProgram.objects.filter(is_active=True).count(),
            'completed_exercises': ExerciseProgress.objects.filter(is_completed=True).count(),
            'survey_responses': SurveyResponse.objects.filter(completed_at__isnull=False).count()
        }
    
    def get_growth_stats(self):
        """Статистика зростання"""
        today = datetime.now().date()
        month_ago = today - timedelta(days=30)
        
        return {
            'new_users_month': User.objects.filter(date_joined__gte=month_ago).count(),
            'new_consultations_month': Consultation.objects.filter(created_at__gte=month_ago).count(),
            'new_programs_month': RehabilitationProgram.objects.filter(created_at__gte=month_ago).count()
        }
    
    def get_performance_stats(self):
        """Статистика продуктивності"""
        return {
            'avg_exercise_score': round(ExerciseProgress.objects.aggregate(avg_score=Avg('score'))['avg_score'] or 0, 1),
            'program_completion_rate': self.calculate_program_completion_rate(),
            'consultation_satisfaction': self.calculate_consultation_satisfaction()
        }
    
    def calculate_program_completion_rate(self):
        """Розрахунок відсотка завершення програм"""
        total_assignments = RehabilitationProgram.objects.aggregate(
            total=Count('assignments')
        )['total'] or 0
        
        completed_assignments = RehabilitationProgram.objects.aggregate(
            completed=Count('assignments', filter=Q(assignments__status='completed'))
        )['completed'] or 0
        
        return round((completed_assignments / total_assignments * 100) if total_assignments > 0 else 0, 1)
    
    def calculate_consultation_satisfaction(self):
        """Розрахунок задоволеності консультаціями"""
        from survey.models import SurveyResponse, QuestionResponse
        from consultation.models import Consultation
        
        try:
            # Отримуємо всі завершені опитування з оцінкою консультацій
            consultation_ratings = QuestionResponse.objects.filter(
                survey_response__completed_at__isnull=False,
                question__question_type='rating',
                question__text__icontains='консультаці',  # Фільтруємо питання про консультації
                rating_answer__isnull=False
            ).values_list('rating_answer', flat=True)
            
            if not consultation_ratings:
                # Якщо немає опитувань, рахуємо на основі статусу консультацій
                completed_consultations = Consultation.objects.filter(
                    status='completed'
                ).count()
                
                total_consultations = Consultation.objects.count()
                
                if total_consultations > 0:
                    # Базовий розрахунок на основі відсотка завершених консультацій
                    completion_rate = (completed_consultations / total_consultations) * 100
                    # Припускаємо, що завершені консультації = задоволеність
                    return min(completion_rate, 100.0)
                else:
                    return 0.0
            
            # Розраховуємо середній рейтинг (переводимо з 1-10 в 0-100)
            avg_rating = sum(consultation_ratings) / len(consultation_ratings)
            satisfaction_percentage = (avg_rating / 10.0) * 100
            
            return round(satisfaction_percentage, 1)
            
        except Exception as e:
            # У разі помилки повертаємо базове значення
            return 75.0


# Заглушки для PDF/CSV/Excel експорту
class ExportPDFView(ExportReportView):
    def get(self, request, report_id):
        report = get_object_or_404(Report, id=report_id)
        return self.export_pdf(report)


class ExportCSVView(ExportReportView):
    def get(self, request, report_id):
        report = get_object_or_404(Report, id=report_id)
        return self.export_csv(report)


class ExportExcelView(ExportReportView):
    def get(self, request, report_id):
        report = get_object_or_404(Report, id=report_id)
        return self.export_excel(report)
