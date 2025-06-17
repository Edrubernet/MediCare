import json
from datetime import datetime
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Count, Avg, Q
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.views import View
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from .models import (
    Survey, Question, QuestionOption, SurveyAssignment, 
    SurveyResponse, QuestionResponse
)
from .serializers import (
    SurveySerializer, SurveyAssignmentSerializer, SurveyResponseSerializer
)
from .forms import SurveyForm, QuestionFormSet, ResponseForm
from patients.models import PatientProfile
from staff.models import User


class SurveyViewSet(viewsets.ModelViewSet):
    serializer_class = SurveySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Лікарі бачать свої опитування, пацієнти - призначені їм
        if self.request.user.user_type == 'staff':
            return Survey.objects.filter(created_by=self.request.user)
        else:
            return Survey.objects.filter(
                assignments__patient__user=self.request.user,
                is_active=True
            ).distinct()
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class SurveyAssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = SurveyAssignmentSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.user_type == 'staff':
            return SurveyAssignment.objects.filter(assigned_by=self.request.user)
        else:
            return SurveyAssignment.objects.filter(patient__user=self.request.user)


class SurveyResponseViewSet(viewsets.ModelViewSet):
    serializer_class = SurveyResponseSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.user_type == 'staff':
            return SurveyResponse.objects.filter(
                assignment__assigned_by=self.request.user
            )
        else:
            return SurveyResponse.objects.filter(
                assignment__patient__user=self.request.user
            )


class SurveyListView(LoginRequiredMixin, ListView):
    """Список опитувань для лікарів"""
    model = Survey
    template_name = 'survey/survey_list.html'
    context_object_name = 'surveys'
    paginate_by = 20
    
    def get_queryset(self):
        if self.request.user.user_type != 'staff':
            return Survey.objects.none()
        
        queryset = Survey.objects.filter(created_by=self.request.user)
        
        # Фільтрація
        active_filter = self.request.GET.get('active')
        if active_filter == 'true':
            queryset = queryset.filter(is_active=True)
        elif active_filter == 'false':
            queryset = queryset.filter(is_active=False)
        
        # Пошук
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | 
                Q(description__icontains=search)
            )
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_filter'] = self.request.GET.get('active', '')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class SurveyCreateView(LoginRequiredMixin, CreateView):
    """Створення нового опитування"""
    model = Survey
    form_class = SurveyForm
    template_name = 'survey/survey_create.html'
    success_url = reverse_lazy('survey:survey_list')
    
    def dispatch(self, request, *args, **kwargs):
        if request.user.user_type != 'staff':
            messages.error(request, _('Only staff members can create surveys.'))
            return redirect('survey:survey_list')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['question_formset'] = QuestionFormSet(self.request.POST)
        else:
            context['question_formset'] = QuestionFormSet()
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        question_formset = context['question_formset']
        
        if not question_formset.is_valid():
            return self.form_invalid(form)
        
        with transaction.atomic():
            form.instance.created_by = self.request.user
            self.object = form.save()
            
            # Зберігаємо питання
            question_formset.instance = self.object
            question_formset.save()
            
            # Зберігаємо варіанти відповідей
            for question_form in question_formset:
                if question_form.cleaned_data and not question_form.cleaned_data.get('DELETE', False):
                    question = question_form.save(commit=False)
                    question.survey = self.object
                    question.save()
                    
                    # Обробляємо варіанти відповідей для питань з вибором
                    if question.question_type in ['single_choice', 'multiple_choice']:
                        options_data = self.request.POST.get(f'question_{question_form.instance.pk}_options', '[]')
                        try:
                            options = json.loads(options_data)
                            for i, option_text in enumerate(options):
                                if option_text.strip():
                                    QuestionOption.objects.create(
                                        question=question,
                                        text=option_text.strip(),
                                        order=i
                                    )
                        except json.JSONDecodeError:
                            pass
        
        messages.success(self.request, _('Survey created successfully.'))
        return HttpResponseRedirect(self.get_success_url())


class SurveyDetailView(LoginRequiredMixin, DetailView):
    """Деталі опитування"""
    model = Survey
    template_name = 'survey/survey_detail.html'
    context_object_name = 'survey'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика призначень
        assignments = self.object.assignments.all()
        context['assignments_stats'] = {
            'total': assignments.count(),
            'pending': assignments.filter(status='pending').count(),
            'completed': assignments.filter(status='completed').count(),
            'expired': assignments.filter(status='expired').count(),
        }
        
        # Останні відповіді
        context['recent_responses'] = SurveyResponse.objects.filter(
            assignment__survey=self.object,
            completed_at__isnull=False
        ).order_by('-completed_at')[:10]
        
        # Питання опитування
        context['questions'] = self.object.questions.all().order_by('order')
        
        return context


class SurveyUpdateView(LoginRequiredMixin, UpdateView):
    """Редагування опитування"""
    model = Survey
    form_class = SurveyForm
    template_name = 'survey/survey_edit.html'
    success_url = reverse_lazy('survey:survey_list')
    
    def dispatch(self, request, *args, **kwargs):
        survey = self.get_object()
        if survey.created_by != request.user:
            messages.error(request, _('You can only edit your own surveys.'))
            return redirect('survey:survey_detail', pk=survey.pk)
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        messages.success(self.request, _('Survey updated successfully.'))
        return super().form_valid(form)


class SurveyDeleteView(LoginRequiredMixin, DeleteView):
    """Видалення опитування"""
    model = Survey
    template_name = 'survey/survey_delete.html'
    success_url = reverse_lazy('survey:survey_list')
    
    def dispatch(self, request, *args, **kwargs):
        survey = self.get_object()
        if survey.created_by != request.user:
            messages.error(request, _('You can only delete your own surveys.'))
            return redirect('survey:survey_detail', pk=survey.pk)
        return super().dispatch(request, *args, **kwargs)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Survey deleted successfully.'))
        return super().delete(request, *args, **kwargs)


class SurveyAssignView(LoginRequiredMixin, TemplateView):
    """Призначення опитування пацієнтам"""
    template_name = 'survey/survey_assign.html'
    
    def dispatch(self, request, *args, **kwargs):
        if request.user.user_type != 'staff':
            messages.error(request, _('Only staff members can assign surveys.'))
            return redirect('survey:survey_list')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['surveys'] = Survey.objects.filter(
            created_by=self.request.user,
            is_active=True
        )
        context['patients'] = PatientProfile.objects.all()
        return context
    
    def post(self, request):
        survey_id = request.POST.get('survey_id')
        patient_ids = request.POST.getlist('patient_ids')
        due_date = request.POST.get('due_date')
        
        if not survey_id or not patient_ids:
            messages.error(request, _('Please select survey and patients.'))
            return redirect('survey:survey_assign')
        
        survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
        
        assigned_count = 0
        for patient_id in patient_ids:
            try:
                patient = PatientProfile.objects.get(id=patient_id)
                
                # Перевіряємо чи не існує вже таке призначення
                existing = SurveyAssignment.objects.filter(
                    survey=survey,
                    patient=patient,
                    status='pending'
                ).exists()
                
                if not existing:
                    SurveyAssignment.objects.create(
                        survey=survey,
                        patient=patient,
                        assigned_by=request.user,
                        due_date=due_date if due_date else None
                    )
                    assigned_count += 1
                    
            except PatientProfile.DoesNotExist:
                continue
        
        if assigned_count > 0:
            messages.success(
                request, 
                _('Survey assigned to %(count)d patients successfully.') % {'count': assigned_count}
            )
        else:
            messages.warning(request, _('No new assignments were created.'))
        
        return redirect('survey:assignment_list')


class SurveyAssignmentListView(LoginRequiredMixin, ListView):
    """Список призначень опитувань"""
    model = SurveyAssignment
    template_name = 'survey/assignment_list.html'
    context_object_name = 'assignments'
    paginate_by = 20
    
    def get_queryset(self):
        if self.request.user.user_type == 'staff':
            queryset = SurveyAssignment.objects.filter(assigned_by=self.request.user)
        else:
            queryset = SurveyAssignment.objects.filter(patient__user=self.request.user)
        
        # Фільтрація за статусом
        status_filter = self.request.GET.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset.select_related(
            'survey', 'patient__user', 'assigned_by'
        ).order_by('-assigned_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_filter'] = self.request.GET.get('status', '')
        context['status_choices'] = SurveyAssignment.STATUS_CHOICES
        return context


class TakeSurveyView(LoginRequiredMixin, TemplateView):
    """Проходження опитування пацієнтом"""
    template_name = 'survey/take_survey.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assignment_id = kwargs.get('assignment_id')
        
        assignment = get_object_or_404(
            SurveyAssignment,
            id=assignment_id,
            patient__user=self.request.user,
            status='pending'
        )
        
        context['assignment'] = assignment
        context['survey'] = assignment.survey
        context['questions'] = assignment.survey.questions.all().order_by('order')
        
        # Перевіряємо чи вже започаткована відповідь
        try:
            response = SurveyResponse.objects.get(assignment=assignment)
            context['existing_response'] = response
            
            # Завантажуємо існуючі відповіді
            existing_answers = {}
            for qr in response.question_responses.all():
                if qr.text_answer:
                    existing_answers[qr.question.id] = qr.text_answer
                elif qr.rating_answer:
                    existing_answers[qr.question.id] = qr.rating_answer
                elif qr.selected_options.exists():
                    existing_answers[qr.question.id] = list(qr.selected_options.values_list('id', flat=True))
            
            context['existing_answers'] = existing_answers
            
        except SurveyResponse.DoesNotExist:
            context['existing_response'] = None
            context['existing_answers'] = {}
        
        return context


class SubmitSurveyView(LoginRequiredMixin, View):
    """Подання відповідей на опитування"""
    
    def post(self, request, assignment_id):
        assignment = get_object_or_404(
            SurveyAssignment,
            id=assignment_id,
            patient__user=request.user,
            status='pending'
        )
        
        try:
            with transaction.atomic():
                # Створюємо або отримуємо відповідь
                response, created = SurveyResponse.objects.get_or_create(
                    assignment=assignment,
                    defaults={'started_at': datetime.now()}
                )
                
                # Очищуємо старі відповіді якщо редагуємо
                if not created:
                    response.question_responses.all().delete()
                
                # Обробляємо відповіді на питання
                questions = assignment.survey.questions.all()
                answered_count = 0
                
                for question in questions:
                    question_response = QuestionResponse.objects.create(
                        survey_response=response,
                        question=question
                    )
                    
                    if question.question_type == 'text':
                        answer = request.POST.get(f'question_{question.id}', '').strip()
                        if answer:
                            question_response.text_answer = answer
                            answered_count += 1
                    
                    elif question.question_type == 'rating':
                        answer = request.POST.get(f'question_{question.id}')
                        if answer:
                            try:
                                rating = int(answer)
                                if 1 <= rating <= 10:
                                    question_response.rating_answer = rating
                                    answered_count += 1
                            except (ValueError, TypeError):
                                pass
                    
                    elif question.question_type == 'yes_no':
                        answer = request.POST.get(f'question_{question.id}')
                        if answer in ['yes', 'no']:
                            question_response.text_answer = answer
                            answered_count += 1
                    
                    elif question.question_type == 'single_choice':
                        option_id = request.POST.get(f'question_{question.id}')
                        if option_id:
                            try:
                                option = QuestionOption.objects.get(
                                    id=option_id,
                                    question=question
                                )
                                question_response.save()
                                question_response.selected_options.add(option)
                                answered_count += 1
                            except QuestionOption.DoesNotExist:
                                pass
                    
                    elif question.question_type == 'multiple_choice':
                        option_ids = request.POST.getlist(f'question_{question.id}')
                        if option_ids:
                            options = QuestionOption.objects.filter(
                                id__in=option_ids,
                                question=question
                            )
                            if options.exists():
                                question_response.save()
                                question_response.selected_options.add(*options)
                                answered_count += 1
                    
                    question_response.save()
                
                # Перевіряємо чи всі обов'язкові питання відповіджено
                required_questions = questions.filter(required=True)
                required_answered = 0
                
                for question in required_questions:
                    qr = response.question_responses.filter(question=question).first()
                    if qr and (qr.text_answer or qr.rating_answer or qr.selected_options.exists()):
                        required_answered += 1
                
                # Якщо не всі обов'язкові питання відповіджено
                if required_answered < required_questions.count():
                    messages.warning(
                        request,
                        _('Please answer all required questions before submitting.')
                    )
                    return redirect('survey:take_survey', assignment_id=assignment_id)
                
                # Завершуємо опитування
                response.completed_at = datetime.now()
                response.save()
                
                # Оновлюємо статус призначення
                assignment.status = 'completed'
                assignment.completed_date = datetime.now()
                assignment.save()
                
                messages.success(request, _('Survey completed successfully. Thank you for your responses!'))
                return redirect('survey:assignment_list')
                
        except Exception as e:
            messages.error(request, _('An error occurred while submitting your responses. Please try again.'))
            return redirect('survey:take_survey', assignment_id=assignment_id)


class SurveyResultsView(LoginRequiredMixin, DetailView):
    """Результати опитування"""
    model = Survey
    template_name = 'survey/survey_results.html'
    context_object_name = 'survey'
    pk_url_kwarg = 'survey_id'
    
    def dispatch(self, request, *args, **kwargs):
        survey = self.get_object()
        if survey.created_by != request.user:
            messages.error(request, _('You can only view results of your own surveys.'))
            return redirect('survey:survey_list')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика опитування
        responses = SurveyResponse.objects.filter(
            assignment__survey=self.object,
            completed_at__isnull=False
        )
        
        context['responses_count'] = responses.count()
        context['assignments_count'] = self.object.assignments.count()
        context['completion_rate'] = round(
            (responses.count() / self.object.assignments.count() * 100) 
            if self.object.assignments.count() > 0 else 0, 1
        )
        
        # Результати по питаннях
        questions_results = []
        for question in self.object.questions.all().order_by('order'):
            question_responses = QuestionResponse.objects.filter(
                survey_response__in=responses,
                question=question
            )
            
            result = {
                'question': question,
                'responses_count': question_responses.count(),
                'results': self.get_question_results(question, question_responses)
            }
            questions_results.append(result)
        
        context['questions_results'] = questions_results
        context['responses'] = responses.order_by('-completed_at')[:20]
        
        return context
    
    def get_question_results(self, question, question_responses):
        """Аналіз результатів конкретного питання"""
        if question.question_type == 'text':
            return {
                'type': 'text',
                'answers': question_responses.values_list('text_answer', flat=True)
            }
        
        elif question.question_type == 'rating':
            ratings = question_responses.filter(
                rating_answer__isnull=False
            ).values_list('rating_answer', flat=True)
            
            if ratings:
                avg_rating = sum(ratings) / len(ratings)
                rating_counts = {}
                for rating in range(1, 11):
                    rating_counts[rating] = list(ratings).count(rating)
                
                return {
                    'type': 'rating',
                    'average': round(avg_rating, 1),
                    'count_by_rating': rating_counts,
                    'total_responses': len(ratings)
                }
            
        elif question.question_type == 'yes_no':
            answers = question_responses.values_list('text_answer', flat=True)
            yes_count = list(answers).count('yes')
            no_count = list(answers).count('no')
            
            return {
                'type': 'yes_no',
                'yes_count': yes_count,
                'no_count': no_count,
                'yes_percentage': round((yes_count / len(answers) * 100) if answers else 0, 1),
                'no_percentage': round((no_count / len(answers) * 100) if answers else 0, 1)
            }
        
        elif question.question_type in ['single_choice', 'multiple_choice']:
            option_counts = {}
            total_selections = 0
            
            for option in question.options.all():
                count = QuestionResponse.objects.filter(
                    survey_response__in=question_responses.values_list('survey_response', flat=True),
                    question=question,
                    selected_options=option
                ).count()
                
                option_counts[option.text] = count
                total_selections += count
            
            # Розрахунок відсотків
            option_percentages = {}
            for option_text, count in option_counts.items():
                percentage = round((count / total_selections * 100) if total_selections > 0 else 0, 1)
                option_percentages[option_text] = percentage
            
            return {
                'type': 'choice',
                'option_counts': option_counts,
                'option_percentages': option_percentages,
                'total_selections': total_selections
            }
        
        return {'type': 'unknown'}


class ExportSurveyResultsView(LoginRequiredMixin, View):
    """Експорт результатів опитування"""
    
    def get(self, request, survey_id):
        survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
        export_format = request.GET.get('format', 'csv')
        
        if export_format == 'csv':
            return self.export_csv(survey)
        elif export_format == 'excel':
            return self.export_excel(survey)
        else:
            messages.error(request, _('Invalid export format.'))
            return redirect('survey:survey_results', survey_id=survey_id)
    
    def export_csv(self, survey):
        """Експорт в CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="survey_{survey.id}_results.csv"'
        
        writer = csv.writer(response)
        
        # Заголовки
        headers = ['Patient', 'Completed Date']
        for question in survey.questions.all().order_by('order'):
            headers.append(question.text)
        writer.writerow(headers)
        
        # Дані
        responses = SurveyResponse.objects.filter(
            assignment__survey=survey,
            completed_at__isnull=False
        ).order_by('-completed_at')
        
        for response in responses:
            row = [
                response.assignment.patient.user.get_full_name(),
                response.completed_at.strftime('%Y-%m-%d %H:%M')
            ]
            
            for question in survey.questions.all().order_by('order'):
                qr = response.question_responses.filter(question=question).first()
                if qr:
                    if qr.text_answer:
                        row.append(qr.text_answer)
                    elif qr.rating_answer:
                        row.append(str(qr.rating_answer))
                    elif qr.selected_options.exists():
                        options = ', '.join(qr.selected_options.values_list('text', flat=True))
                        row.append(options)
                    else:
                        row.append('')
                else:
                    row.append('')
            
            writer.writerow(row)
        
        return response
    
    def export_excel(self, survey):
        """Експорт в Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse
        import io
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Survey Results"
        
        # Заголовки
        headers = ['Patient', 'Completed Date']
        for question in survey.questions.all().order_by('order'):
            headers.append(question.text)
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Дані
        responses = SurveyResponse.objects.filter(
            assignment__survey=survey,
            completed_at__isnull=False
        ).order_by('-completed_at')
        
        for row_num, response in enumerate(responses, 2):
            worksheet.cell(row=row_num, column=1, value=response.assignment.patient.user.get_full_name())
            worksheet.cell(row=row_num, column=2, value=response.completed_at.strftime('%Y-%m-%d %H:%M'))
            
            col_num = 3
            for question in survey.questions.all().order_by('order'):
                qr = response.question_responses.filter(question=question).first()
                if qr:
                    if qr.text_answer:
                        worksheet.cell(row=row_num, column=col_num, value=qr.text_answer)
                    elif qr.rating_answer:
                        worksheet.cell(row=row_num, column=col_num, value=qr.rating_answer)
                    elif qr.selected_options.exists():
                        options = ', '.join(qr.selected_options.values_list('text', flat=True))
                        worksheet.cell(row=row_num, column=col_num, value=options)
                col_num += 1
        
        # Зберігаємо в пам'ять
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="survey_{survey.id}_results.xlsx"'
        
        return response


# API для динамічної взаємодії
class SurveyStatsAPIView(LoginRequiredMixin, View):
    """API для статистики опитувань"""
    
    def get(self, request, survey_id):
        survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
        
        assignments = survey.assignments.all()
        responses = SurveyResponse.objects.filter(
            assignment__survey=survey,
            completed_at__isnull=False
        )
        
        stats = {
            'total_assignments': assignments.count(),
            'completed_responses': responses.count(),
            'pending_assignments': assignments.filter(status='pending').count(),
            'completion_rate': round(
                (responses.count() / assignments.count() * 100) 
                if assignments.count() > 0 else 0, 1
            ),
            'average_completion_time': self.calculate_avg_completion_time(responses),
            'recent_completions': responses.order_by('-completed_at')[:5].values(
                'assignment__patient__user__first_name',
                'assignment__patient__user__last_name',
                'completed_at'
            )
        }
        
        return JsonResponse(stats)
    
    def calculate_avg_completion_time(self, responses):
        """Розрахунок середнього часу виконання"""
        completion_times = []
        for response in responses:
            if response.started_at and response.completed_at:
                duration = response.completed_at - response.started_at
                completion_times.append(duration.total_seconds() / 60)  # в хвилинах
        
        if completion_times:
            return round(sum(completion_times) / len(completion_times), 1)
        return 0
